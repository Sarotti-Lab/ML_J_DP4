# -*- coding: utf-8 -*-
"""
Created on Tue Mar 15 2022
@author: María M. Zanardi & Ariel M. Sarotti

-----------------------------------------------------------------------------------
-----------------------------------------------------------------------------------
          This program lets you to perform ML-J DP4 calculation automatically.
The program allows the use of H, C and J data. Although it can give partial results
using some subset of data, it is recommended to use the complete data set.

*Conditions for its use:

        You must place in a folder: all the outputs containing NMR and NBO
calculations, corresponding to all posible conformations and isomers under
analysis (in format *.log or *.out). The names of the file will necessarily
be coded as IsomerNumber_*.

        Additionally the folder must also contain an excel file consigning the 
experimental data an labels ordered as follows:
    -experimental coupling constants and Labels of the Hs coupled (sheet name = 'J')
    -experimental chemical shifts and labels of the atoms (sheet namme = 'shifts')
-------------------------------------------------------------------------------------
-------------------------------------------------------------------------------------
"""

__version__ = "1.3.2"
__author__ = 'Ariel M. Sarotti & María M. Zanardi'

import glob
import pickle
import os
from tkinter import Tk, filedialog
import pandas as pd
import numpy as np
import math
from math import isnan
from math import cos
from sklearn import linear_model
import scipy.stats as stats
import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import re


def change_directory(s):
    'Define work directory'
    print(f'\n{s:s}')
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    di = filedialog.askdirectory()
    if di:
        os.chdir(di)
        print(f' -> Using the directory \n    {di:s}')
        return True
    else:
        print(' -> cancel.')
        return False
        

def data_sheet(cant_comp):
    'Allows you to read Excel with the label, interactively \ n'
    print('\nSelect the excel file with the experimental data and labels.')
    open_file = filedialog.askopenfilename()
    if open_file=='':
        print(' -> cancel.')
        return False

        
        
    print(f' -> Using {open_file:s} as excel file with experimental data and labels.')
    dg = pd.read_excel(open_file, sheet_name='J',engine='openpyxl')
    couplings = np.array(dg[dg['exp'].isna() == False])
    
    J_exp = pd.DataFrame([couplings[i][0:2] for i in range(couplings.shape[0])])
    if couplings.shape[1] < 5:
        Jwtl = np.array([couplings[i][2:4] for i in range(couplings.shape[0])])
    else:
        for i in range(cant_comp):
            end_label = (cant_comp *2) + 3
            Jwtl = np.array([couplings[i][2:end_label] for i in range(couplings.shape[0])])
    Jwtl = Jwtl.astype(int)

    df = pd.read_excel(open_file, sheet_name='shifts',engine='openpyxl')
    shifts = np.array(df[df['nuclei'].isna() == False])
    d_exp_C = np.array([shifts[i][1:3] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
    d_exp_H = np.array([shifts[i][1:3] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])
    if shifts.shape[1] < 7:
        wtl_C = np.array([shifts[i][3:6] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
        wtl_H = np.array([shifts[i][3:6] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])
    else:
        for i in range(cant_comp):
            end_label = (cant_comp *3) + 3
            wtl_C = np.array([shifts[i][3:end_label] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
            wtl_H = np.array([shifts[i][3:end_label] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])

    return J_exp, Jwtl, d_exp_C, wtl_C, d_exp_H, wtl_H

def label_check(wtl, isom):
    try:
        if len(wtl[0]) < 4:
            return wtl
        else:
            start = (isom -1) *3
            end = start + 3
            return wtl[:, start:end]
    except:
        return wtl

def label_check_J(Jwtl, isom):
    try:
        if len(Jwtl[0]) < 4:
            return Jwtl
        else:
            start = (isom -1) *2
            end = start + 2
            return Jwtl[:, start:end]
    except:
        return Jwtl

def isomer_count():
    'Employs de filename to diferenciate and count the number of isomers'
    lista= glob.glob('*.log') + glob.glob('*.out')
    isomer_list =[]
    for e in lista:
        if e[0:3] not in isomer_list:
            isomer_list.append(e[0:3])
        else:
            continue
    isomer_list.sort() ##RG
    isomer_list.sort(key=lambda s: len(s)) #RG    
    return len(isomer_list), isomer_list

def get_energy(file):
    'Extract SCF energies of every Gaussian 09 output'
    with open (file,'rt') as f:
        lines=f.readlines()
        for i, line in enumerate(lines):
            if "SCF Done:" in line:
                energy=float(line.split()[4])
    return energy

def relative_energies(energies):
    'Receive a list with the energies of all the conformers and get the energy relative to the minimum'
    energ = np.array(energies)
    energ *= 627.5095
    mas_estable = energ.min()
    e_relativas = energ - mas_estable
    return e_relativas

def get_information(file):
    '''Recive an output of gaussian 09 (.log o .out) and return 10 elements:
    The txt file in lines, SCF energy, a tuple of tensors, the start and final
    index of the following matrices: Coordinate matrix, Distance matrix,
    Mulliken matrix and NBO analysis.
    '''
    tensors=[]
    start_mat = None
    end_mat = None

    with open (file,'rt') as f:
        lines=f.readlines()
        for i, line in enumerate(lines):
            if "SCF Done:" in line:
                energy=float(line.split()[4])
            if "Isotropic = " in line:
                tensors.append(float(line.split()[4]))
            if "Distance matrix" in line:
                start_mat= i
            if "Stoichiometry" in line:
                end_mat = i
            if 'Standard orientation:' in line:
                start_coord = i + 4
            if 'Rotational constants' in line:
                end_coord = i-1
            if " Summary of Natural Population Analysis" in line:
                start_mull = i + 5
            if "Bond orbital/ Coefficients/ Hybrids" in line:
                start_nbo = i
            if "NHO Directionality and" in line:
                end_nbo = i

    return lines, energy, tuple(tensors), start_coord, end_coord, start_mat, end_mat, start_mull, start_nbo, end_nbo

def matrix(lines, start_coord, end_coord, start_mat, end_mat, start_mull):
    '''This function receives the txt and extracts molecular size, the
    coordinate matrix, the distance matrix and the mulliken vector
    '''
    mol_size = end_coord - start_coord -1
    coord_mat=[] 
    dist_mat = []
    mull_mat = []

    for i, line in enumerate(lines):
        
        if i > start_coord and i< end_coord:
            row = line.split()[:]
            row = [float(_) for _ in row]
            coord_mat.append(row) 

        if start_mat!= None and i > start_mat and i < end_mat:
            dist_mat.append(line.split()[2:])
        if i > start_mull and i < (start_mull + mol_size + 1):
            mull_mat.append(float(line.split()[2]))

        if i >= (start_mull + mol_size):
            break
    coord_mat = np.array(coord_mat)

    if start_mat == None:
        dist_mat = dist_mat_calculator(coord_mat)
    else:
        dist_mat= dist_mat_format(dist_mat, mol_size)

    return mol_size, coord_mat, dist_mat, mull_mat

def dist_mat_calculator(coord_mat):
    coordenates = coord_mat[:,3:6]
    dist_mat = np.zeros([len(coordenates), len(coordenates)])
    for i in range(dist_mat.shape[0]):
        for j in range(i):
            dist_mat[i][j] = np.sqrt(sum(np.power((coordenates[i]-coordenates[j]),2)))

    return dist_mat

def dist_mat_format(dist_mat, mol_size):
    '''This function formats the Gaussian distance matrix as a triangular matrix'''

    lotes = math.ceil(mol_size/5)
    dist_mat.pop(0)
    eliminar = mol_size
    largo_lote = mol_size
    for lote in range(lotes-1):
        dist_mat.pop(eliminar)
        largo_lote -= 5
        eliminar += largo_lote

    new_mat = np.zeros((mol_size, mol_size))
    n = 0
    s = 0
    row_index = 0
    col_index = 0
    for lote in range(lotes):
        for x in range(s, mol_size):
            col_index = 0
            for y in range(n, n+5):
                try:
                    new_mat[x][y]= float(dist_mat[row_index][col_index])
                    col_index += 1
                except:
                    continue
            row_index += 1
        n += 5
        s += 5
    return new_mat

def buscar_vecinos(dist_mat, coord_mat):
    '''Recive a distance and coordinate matrix and return a matrix of 0 and 1
    depending on whether or not they are neighbors'''
    pesos = [coord_mat[i][1] for i in range(len(coord_mat))]
    mat_vecinos = dist_mat.copy()
    atoms = len(pesos)
    for col in range(atoms):
        for row in range(atoms):
            a = dist_mat[row][col]
            if a < 2.1 and a > 0 and (pesos[row] > 18 or pesos[col] > 18):
                mat_vecinos[row][col] = 1
            elif a < 1.9 and a > 0 and (pesos[row]< 18 or pesos[col]< 18):
                mat_vecinos[row][col] = 1
            else:
                mat_vecinos[row][col] = 0
    mat_transpuesta = np.transpose(mat_vecinos)
    vecinos = mat_vecinos + mat_transpuesta
    return vecinos

def hibridizations(lines, mol_size, start_nbo, end_nbo):
    '''It allows obtaining an  hybridizations matrix of the
    neighbors to each nucleus by reading nbo information'''
    hibridization_mat = np.zeros((mol_size, mol_size))
    for i in range(start_nbo, end_nbo):
        if "BD" in lines[i]:
            n=0
            j=0
            row = [c for c in re.split(' |-',lines[i]) if c!='']
            if float(row[1][1:5]) > 1:
                index_x = int(row[6])-1
                index_y = int(row[8])-1
                while n < 2:
                    j += 1
                    if '%)' in lines[i+j] and n == 0:
                        n += 1
                        hib = lines[i+j]
                        a = [n for n,e in enumerate(hib) if e == '%']
                        hibrid_x = float(hib[a[1]-6 : a[1]-1])
                    elif '%)' in lines[i+j] and n == 1 and row[7] == 'H':
                        n += 1
                        hibrid_y = 100
                    elif '%)' in lines[i+j] and n == 1 and row[7] != 'H':
                        n += 1
                        hib = lines[i+j]
                        a = [n for n,e in enumerate(hib) if e == '%']
                        hibrid_y = float(hib[a[1]-6 : a[1]-1])
                    else:
                        continue
                            
            if hibrid_x > 10:
                hibridization_mat[index_x][index_y] = hibrid_x/100
                hibridization_mat[index_y][index_x] = hibrid_y/100
            else:
                continue
        else:
            continue
    return hibridization_mat

def buscar_vecinos_H(hibridization_mat):
    vecinos = []
    for row in range(hibridization_mat.shape[0]):
        vec =[]
        for n, col in enumerate(hibridization_mat[row]):
            if col != 0:
                vec.append(n+1)
        vecinos.append(vec)
    vecinos = pd.DataFrame(vecinos)
    return vecinos

def energy_check(energies,energy):
    '''This function gives a warning to the user about settling for repeated
    energies considering both in the analysis. The user is responsible for
    eliminating one of them if they are repeated conformations.
    '''
    if energy in energies:
        print (f'\nUds tiene compuestos con energías {energy} repetidas')
        energy += np.random.randint(10,100)/10**10
        print(f'Uno de ellos se reemplazo por {energy}')
    return energy

def indexation(coord_mat):
    index_C = []
    index_H = []
    for i, e in enumerate(coord_mat):
        if e[1] == 6:
            index_C.append(i)
        elif e[1] == 1:
            index_H.append(i)
        else:
            continue
    return np.array(index_C), np.array(index_H)

def descriptor_calculator(tensors, coord_mat, mull_mat, hibridization_mat, mat_vecinos, dist_mat):
    '''Function that receives a series of matrices extracted from the Gaussian
    output and calculates a series of molecular descriptors through the call
    to other more specific functions
    '''
    index_C, index_H = indexation(coord_mat)
    desc_mat_C = np.zeros((len(index_C), 114))
    desc_mat_H = np.zeros((len(index_H), 114))

    carga_v = carga_vecinos(mull_mat, hibridization_mat)
    n_at_v = n_at_vecinos(coord_mat, hibridization_mat)
    hib_v = hib_vecinos(hibridization_mat, mat_vecinos)
    dist_v = dist_vecinos(hibridization_mat, dist_mat)
    indices_vecinos = vec_index(hibridization_mat)
    angulos_vecinos = angulos(coord_mat, indices_vecinos)
    matriz_culomb = culomb(dist_mat, coord_mat)
    matriz_tensorial_1, matriz_tensorial_2, matriz_tensorial_1_H, matriz_tensorial_2_H = tensorial(dist_mat, tensors, coord_mat)

    for i, e in enumerate(index_C):
        desc_mat_C[i][0] = tensors[e] 
        desc_mat_C[i][1] = mull_mat[e]  
        desc_mat_C[i][2:6] = carga_v[e] 
        desc_mat_C[i][6:10] = n_at_v [e] 
        desc_mat_C[i][10:14] = hib_v [e] 
        desc_mat_C[i][14:18] = dist_v [e] 
        desc_mat_C[i][18:24] = angulos_vecinos [e] 
        desc_mat_C[i][24:54] = matriz_culomb [e][0:30]
        desc_mat_C[i][54:84] = matriz_tensorial_1 [e][0:30] 
        desc_mat_C[i][84:114] = matriz_tensorial_2 [e][0:30] 

    for i, e in enumerate(index_H):
            index_C_unido = hibridization_mat[e].nonzero()[0][0]
            desc_mat_H[i][0] = tensors[e]
            desc_mat_H[i][1] = mull_mat[index_C_unido]
            desc_mat_H[i][2:6] = carga_v[index_C_unido] 
            desc_mat_H[i][6:10] = n_at_v [index_C_unido] 
            desc_mat_H[i][10:14] = hib_v [index_C_unido] 
            desc_mat_H[i][14:18] = dist_v [index_C_unido] 
            desc_mat_H[i][18:24] = angulos_vecinos [index_C_unido] 
            desc_mat_H[i][24:54] = matriz_culomb [e][1:31]
            desc_mat_H[i][54:84] = matriz_tensorial_1_H [e][1:31]
            desc_mat_H[i][84:114] = matriz_tensorial_2_H [e][1:31]

    return desc_mat_C, desc_mat_H, index_C, index_H

def carga_vecinos(mull_mat, hibridization_mat):
    '''Employing the hybridization and the Mulliken matrix it let you obtain
    the charge of the atoms attached to the atom analized'''
    carga_vecinos = []
    for i in range(hibridization_mat.shape[0]):
        q_vecinos = []
        k = 0
        for j in range(hibridization_mat.shape[1]):
            if hibridization_mat[i][j] != 0:
                q_vecinos.append(mull_mat[j])
                k += 1
            else:
                continue
        q_vecinos.sort(reverse= True)
        if len(q_vecinos) < 4:
            q_vecinos += [0] * (4 - len(q_vecinos))

        carga_vecinos.append(q_vecinos)
    carga_vecinos = np.array(carga_vecinos)
    return carga_vecinos

def n_at_vecinos(coord_mat, hibridization_mat):
    n_at_vecinos = np.zeros((len(coord_mat), 4))
    for i in range(hibridization_mat.shape[0]):
        k = 0
        for j in range(hibridization_mat.shape[1]):
            if hibridization_mat[i][j] != 0:
                n_at_vecinos[i][k] = coord_mat[j][1]
                k += 1
            else:
                continue
        n_at_vecinos[i] = np.sort(n_at_vecinos[i])[::-1]
    return n_at_vecinos

def hib_vecinos(hibridization_mat, mat_vecinos):
    hib_vecinos = []
    for i in range(mat_vecinos.shape[0]):
        k = 0
        hyb =[]
        for j in range(mat_vecinos.shape[1]):
            if hibridization_mat[i][j] != 0:
                hyb.append(hibridization_mat[j][i])
                k += 1
            else:
                continue
        hyb.sort()
        if len(hyb) < 4:
            hyb += [0] * (4 - len(hyb))

        hib_vecinos.append(hyb)
    hib_vecinos = np.array(hib_vecinos)
    return hib_vecinos

def dist_vecinos(hibridization_mat, dist_mat):
    '''It allows you obtain a matrix with the hibridization of all the atoms
    binding to the nuecleus up to 4 atoms for C sp3'''
    dist_mat = dist_mat + dist_mat.transpose()
    dist_vecinos = []
    for i in range(hibridization_mat.shape[0]):
        k = 0
        dist =[]
        for j in range(hibridization_mat.shape[1]):
            if hibridization_mat[i][j] != 0:
                dist.append(dist_mat[i][j])
                k += 1
            else:
                continue
        dist.sort(reverse=True)
        if len(dist) < 4:
            dist += [0] * (4 - len(dist))

        dist_vecinos.append(dist)
    dist_vecinos = np.array(dist_vecinos)
    return dist_vecinos

def vec_index(hibridization_mat):
    '''Function that allows to index the neighbors of a certain nucleus'''
    indices_vecinos = []
    for i in range(hibridization_mat.shape[0]):
        indices =[]
        for j in range(hibridization_mat.shape[1]):
            if hibridization_mat[i][j] != 0:
                indices.append(j)
            else:
                continue
        #if len(indices) < 4:
            #hyb += [0] * (4 - len(indices))
        indices_vecinos.append(indices)
    return indices_vecinos

def angulos(coord_mat, indices_vecinos):
    '''Allows you to calculate the angles that form different atomic triads'''
    angulos_vecinos = []
    for i in range(len(indices_vecinos)):
        punto_central = coord_mat[i][3:6] 
        coord_vecinos=[]
        for j in range(4):
            try:
                p = coord_mat[indices_vecinos[i][j]][3:6]
            except:
                p = [0,0,0]
            coord_vecinos.append(p)
        coord_vecinos = np.array(coord_vecinos)
        vectores = []
        for vecino in range(len(coord_vecinos)):
            if sum(coord_vecinos[vecino]) !=0:
                vector = (coord_vecinos[vecino] - punto_central)
            else:
                vector = coord_vecinos[vecino]
            vectores.append(vector)
        vectores = np.array(vectores)
        cuadrado_vectores = vectores**2
        v_norm = vectores.copy()
        for e in range(4):
            v_mag = np.sqrt(sum(cuadrado_vectores[e]))
            if v_mag != 0:
                v_norm[e] = vectores[e]/v_mag
            else:
                v_norm[e] = vectores[e]
        res = np.zeros(6)
        res[0] = sum(v_norm[0][i]*v_norm[1][i] for i in range(3))
        res[1] = sum(v_norm[0][i]*v_norm[2][i] for i in range(3))
        res[2] = sum(v_norm[0][i]*v_norm[3][i] for i in range(3))
        res[3] = sum(v_norm[1][i]*v_norm[2][i] for i in range(3))
        res[4] = sum(v_norm[1][i]*v_norm[3][i] for i in range(3))
        res[5] = sum(v_norm[2][i]*v_norm[3][i] for i in range(3))
        angle_rad = np.arccos(res)
        angulos_grados = [math.degrees(angle_rad[i]) if math.degrees(angle_rad[i]) != 90 else 0 for i in range(6)]
        angulos = sorted([cos(angle_rad[i]) for i in range(6) if angulos_grados[i] > 90])  
        angulos = angulos + [0 for _ in range(6 - len(angulos))]
        angulos_vecinos.append(angulos)

    angulos = np.array(angulos_vecinos)

    return angulos

def culomb(dist_mat, coord_mat):
    matriz_culomb = np.zeros((dist_mat.shape[0], dist_mat.shape[1]))
    dist_mat_cuadrada = dist_mat + dist_mat.transpose()
    for i in range(dist_mat.shape[0]):
        for j in range(dist_mat.shape[1]):
            if dist_mat_cuadrada[i][j]>0 and dist_mat_cuadrada[i][j] < 5:
                if coord_mat[i][1] > 1:
                    matriz_culomb[i][j] = ((coord_mat[i][1] * coord_mat[j][1]) / dist_mat_cuadrada[i][j])/150
                   
                else:
                    matriz_culomb[i][j] = ((coord_mat[i][1] * coord_mat[j][1]) / dist_mat_cuadrada[i][j])/17 
            else:
                matriz_culomb[i][j] = 0
        matriz_culomb[i] = np.sort(matriz_culomb[i])[::-1]
    if matriz_culomb.shape[1] < 31:
        zeros = np.zeros([matriz_culomb.shape[0],(31-matriz_culomb.shape[1])])
        matriz_culomb = np.append(matriz_culomb, zeros, axis =1)
    return matriz_culomb

def tensorial(dist_mat, tensors, coord_mat):
    '''Performs the calculation of the tensorial matrix by divided the tensor
    of the different nuclei in a radius of 5 A and the distance and the square
    or cube of the distance as coorespond'''
    np.seterr(divide='ignore', invalid='ignore')
    distancias = dist_mat.copy()
    distancias[distancias > 5]= 1000 
    tensores = np.array(tensors)
    escalado = 2.8 * 150
    escalado_2 = 1.5 * 150
    escalado_H = 10*17
    escalado_H2 = 4*17

    distancias_cuadrada = np.power((distancias + distancias.transpose()), 2)
    distancias_cubo = np.power((distancias + distancias.transpose()), 3)
    mat_ones_1 = np.ones((tensores.shape[0],tensores.shape[0]))
    mat_ones_2 = np.ones((tensores.shape[0],tensores.shape[0]))
    matriz_1=(mat_ones_1*abs(tensores)/distancias_cuadrada)
    matriz_2=(mat_ones_2*abs(tensores)/distancias_cubo)

    matriz_1_C= matriz_1 / escalado * -1
    matriz_2_C= matriz_2 / escalado_2 * -1
    matriz_1_H= matriz_1 / escalado_H * -1
    matriz_2_H= matriz_2 / escalado_H2 * -1
    matriz_1_C.sort(axis=1)
    matriz_2_C.sort(axis=1)
    matriz_1_H.sort(axis=1)
    matriz_2_H.sort(axis=1)

    matriz_tensorial_1 = abs(np.delete(matriz_1_C, 0, axis= 1))
    matriz_tensorial_2 = abs(np.delete(matriz_2_C, 0, axis= 1))
    matriz_tensorial_1_H = abs(np.delete(matriz_1_H, 0, axis= 1))
    matriz_tensorial_2_H = abs(np.delete(matriz_2_H, 0, axis= 1))


    if matriz_tensorial_1.shape[1] < 31:
        zeros_C = np.zeros([tensores.shape[0],(31-matriz_tensorial_1.shape[1])])
        matriz_tensorial_1 = np.append(matriz_tensorial_1, zeros_C, axis=1)
        matriz_tensorial_2 = np.append(matriz_tensorial_2, zeros_C, axis=1)


    if matriz_tensorial_1_H.shape[1] < 31:
        zeros_H = np.zeros([tensores.shape[0],(31-matriz_tensorial_1_H.shape[1])])
        matriz_tensorial_1_H = np.append(matriz_tensorial_1_H, zeros_H, axis=1)
        matriz_tensorial_2_H = np.append(matriz_tensorial_2_H, zeros_C, axis=1)

    return matriz_tensorial_1, matriz_tensorial_2, matriz_tensorial_1_H, matriz_tensorial_2_H

def carga_kernels():
    kr_dir = Path(__file__).parent / "data"
    krC = pickle.load(open((kr_dir / 'krC.dat').as_posix(), 'rb'))
    krH = pickle.load(open((kr_dir / 'krH.dat').as_posix(), 'rb'))
    return krC, krH

def tensors_correction(descriptors_C, descriptors_H, krC, krH):
    '''Employs the trained kernels to adjust the low level tensors to a tensor
    obtained at a better level by using the descriptors calculated formerly'''
    tensors_corrected_C = krC.predict(descriptors_C)
    tensors_corrected_H = krH.predict(descriptors_H)
    return tensors_corrected_C, tensors_corrected_H

def tensor_indexation(tensors_corrected_C, tensors_corrected_H, index_C, index_H, mol_size):
    '''This function let index the tensors to be able to order them according
    to the labels'''
    tensors_indexed_C = dict(zip(index_C, tensors_corrected_C))
    tensors_indexed_H = dict(zip(index_H, tensors_corrected_H))
    return tensors_indexed_C, tensors_indexed_H

def Boltzman(tens_all_conf_C, tens_all_conf_H, jotas_all_conf, e_relativas):
    '''Once all the tensors and J have been extracted and corrected, this function
     performs the weighted averages according to the Boltzman probability.
     Its parameters require a DataFrame with the tensors of C, H and J
     and a list of those relative energies.
    '''

    P_Boltz = np.exp(-e_relativas*4.18/2.5) 
                                        
    contribucion = P_Boltz / P_Boltz.sum()

    tensores_ponderados_C = pd.DataFrame(tens_all_conf_C) * contribucion
    tensores_ponderados_C = tensores_ponderados_C.sum(axis=1)
    tensores_ponderados_H = pd.DataFrame(tens_all_conf_H) * contribucion
    tensores_ponderados_H = tensores_ponderados_H.sum(axis=1)
    try:
        jotas_ponderadas = pd.DataFrame(jotas_all_conf) * contribucion
        jotas_ponderadas = jotas_ponderadas.sum(axis=1)
    except:
        jotas_ponderadas = pd.DataFrame()
    return tensores_ponderados_C, tensores_ponderados_H, jotas_ponderadas

def tens_ordenados(tensors_indexed_C, tensors_indexed_H, wtl_C, wtl_H):
    ''' To operate you need the tensors indexed in Gaussian order and the
    labels of the compound in np.array format (this should have only 3 columns)
    The result is a list with the tensors ordered according to the inserted
    label'''
    tens_C = np.zeros(wtl_C.shape)
    tens_H = np.zeros(wtl_H.shape)
    for i in range (wtl_C.shape[0]):
        for j in range (3):
            if not isnan(wtl_C[i,j]):
                index = int(wtl_C[i,j])
                tens_C[i,j] = tensors_indexed_C[index-1] 
            else:
                tens_C[i,j]=float('nan')
    for i in range (wtl_H.shape[0]):
        for j in range (3):
            if not isnan(wtl_H[i,j]):
                index = int(wtl_H[i,j])
                tens_H[i,j] = tensors_indexed_H[index-1]
            else:
                tens_H[i,j]=float('nan')
    tens_C = pd.DataFrame(tens_C) 
    tens_C = tens_C.mean(axis=1)
 
    tens_H = pd.DataFrame(tens_H) 
    tens_H = tens_H.mean(axis=1)
    return tens_C, tens_H

def d_calculation(tensors_C, tensors_H):
    '''It use the tensors for the calculation of the chemical shifts using
    TMS as a reference standard'''
    TMS_C = 192.29325 
    TMS_H = 31.83057
    Unscaled_shift_C = TMS_C - tensors_C
    Unscaled_shift_H = TMS_H - tensors_H
    return Unscaled_shift_C, Unscaled_shift_H


def escalado_CyH(Unscaled_shift, exp):
    '''Performs C and H scaling.
     You will also order interchangeable H or C but it is necessary that the
     Experimental chemical shifts are ordered from highest to lowest, and
     indicating in the excel with 1 and 0'''
    shifts = []
    exchange = []

    if exp.shape[0] != 0:
        for i in range(exp.shape[0]):
            shifts.append(exp[i][0])
            exchange.append(exp[i][1])
    else:
        return pd.DataFrame(exp), pd.DataFrame(exp)

    UnsC_d = Unscaled_shift.copy()
    indices_intercambiables = [i for i in range(len(exchange)) if exchange[i] ==1]
    for e in indices_intercambiables:
        UnsC_d[e] = max(Unscaled_shift[e], Unscaled_shift[e+1])
        UnsC_d[e+1] = min(Unscaled_shift[e], Unscaled_shift[e+1])


    shifts = np.array(shifts).reshape(-1,1)
    UnsC_d = np.array(UnsC_d).reshape(-1,1)
    regresion = linear_model.LinearRegression()

    regresion.fit(shifts, UnsC_d) 
   
    m = regresion.coef_       
    b = regresion.intercept_    

    Scaled_shift = (UnsC_d - b) / m
    Scaled_shift = pd.DataFrame(Scaled_shift)

    shifts = pd.DataFrame(shifts)
    return Scaled_shift, shifts

def crossproduct(v1, v2):

    product = [0, 0, 0]
    product[0] = v1[1]*v2[2]-v1[2]*v2[1]
    product[1] = v1[2]*v2[0]-v1[0]*v2[2]
    product[2] = v1[0]*v2[1]-v1[1]*v2[0]
    return product

def dotproduct(v1, v2):
    return sum((a*b) for a, b in zip(v1, v2))

def J_label_converter(Jwtl, mat_vecinos_H):
    J_labels = []
    for i in range(Jwtl.shape[0]):
        C_bonded = [Jwtl[i][0]]
        for n, atom in enumerate(Jwtl[i]):
            C_bonded.append(mat_vecinos_H[0][atom-1])
        C_bonded.append(Jwtl[i][1])
        J_labels.append(C_bonded)
    Jwtl = np.array(J_labels)
    return Jwtl

def J_calculator(Jwtl, coord_mat, hibridization_mat):
    '''Calculate the dihedral angle formed by 4 points in space using the
    coordinate matrix of the 4 atoms to be analyzed. Then it use this angle to
    calculate the scalar coupling employing the ecuation of Haasnoot Leeuw Altona
    '''
    ENs = [2.20,0.00,0.98,1.57,2.04,2.6,3.05,3.5,3.9,0.00,
       0.93,1.31,1.61,1.90,2.19,2.6,3.15,0.00,0.82,1.00,
       1.36,1.54,1.63,1.66,1.55,1.83,1.88,1.91,1.90,1.65,
       1.81,2.01,2.18,2.55,2.95,3.00,0.82,0.95,1.22,1.33,
       1.60,2.16,1.90,2.20,2.28,2.20,1.93,1.69,1.78,1.96,
       2.05,2.10,2.66]

    PARAM = [13.86, -0.81, 0, 0.56, -2.32, 17.9]

    index_vec = vec_index(hibridization_mat)

    Jotas = []

    for i in range(Jwtl.shape[0]):
        J_mat = np.zeros((4,3))
        indices_vecinos = copy.deepcopy(index_vec)
        for j, n in enumerate(Jwtl[i]):
            J_mat[j]  = coord_mat[n-1][3:6]

        
        vecinos_C1 = indices_vecinos[Jwtl[i][1]-1] 
        vecinos_C2 = indices_vecinos[Jwtl[i][2]-1]

        for k in [0,2]: 
            vecinos_C1.remove(Jwtl[i][k]-1)
            vecinos_C2.remove(Jwtl[i][k+1]-1)

        EN_C1 = [ENs[elem-1] for elem in [int(coord_mat[i][1]) for i in vecinos_C1]]
        EN_C2 = [ENs[elem-1] for elem in [int(coord_mat[i][1]) for i in vecinos_C2]]

        DelEN1a = EN_C1[0]-2.2
        DelEN2a = EN_C1[1]-2.2
        DelEN1b = EN_C2[0]-2.2
        DelEN2b = EN_C2[1]-2.2

    
        q1 = np.subtract(J_mat[1],J_mat[0])
        q2 = np.subtract(J_mat[2],J_mat[1])
        q3 = np.subtract(J_mat[3],J_mat[2])

        q1_x_q2=np.cross(q1,q2)
        q2_x_q3=np.cross(q2,q3)

        n1=q1_x_q2/np.sqrt(np.dot(q1_x_q2,q1_x_q2))
        n2=q2_x_q3/np.sqrt(np.dot(q2_x_q3,q2_x_q3))

        u1 = n2
        u3 = q2/(np.sqrt(np.dot(q2,q2)))
        u2 = np.cross(u3,u1)

        cos_theta = np.dot(n1,u1)
        sin_theta = np.dot(n1,u2)

        theta = abs(-math.atan2(sin_theta,cos_theta))
        theta_deg = abs(np.degrees(theta))



        mat_coord_vecC1 = np.array([coord_mat[v][3:6] for v in vecinos_C1])
        mat_coord_vecC2 = np.array([coord_mat[v][3:6] for v in vecinos_C2])

        Bvect1a = mat_coord_vecC1[0] - J_mat[1]
        Bvect2a = mat_coord_vecC1[1] - J_mat[1]
        Bvectab = J_mat[2] - J_mat[1]

        Bvect1b = mat_coord_vecC2[0] - J_mat[2]
        Bvect2b = mat_coord_vecC2[1] - J_mat[2]
        Bvectba = J_mat[1] - J_mat[2]


        crossproda = crossproduct(Bvect1a, Bvect2a)
        crossprodb = crossproduct(Bvect1b, Bvect2b)

        signa=np.sign(dotproduct(crossproda, Bvectab))
        if signa==1:
            signa=[1,-1]
        else:
            signa=[-1,1]
        signb=np.sign(dotproduct(crossprodb, Bvectba))
        if signb==1:
            signb=[1,-1]
        else:
            signb=[-1,1]

        Corr1=DelEN1a*(PARAM[3]+PARAM[4]*np.power(np.cos(np.radians(signa[0]*theta_deg+PARAM[5]*np.abs(DelEN1a))),2))
        Corr2=DelEN2a*(PARAM[3]+PARAM[4]*np.power(np.cos(np.radians(signa[1]*theta_deg+PARAM[5]*np.abs(DelEN2a))),2))
        Corr3=DelEN1b*(PARAM[3]+PARAM[4]*np.power(np.cos(np.radians(signb[0]*theta_deg+PARAM[5]*np.abs(DelEN1b))),2))
        Corr4=DelEN2b*(PARAM[3]+PARAM[4]*np.power(np.cos(np.radians(signb[1]*theta_deg+PARAM[5]*np.abs(DelEN2b))),2))

        Corr=Corr1+Corr2+Corr3+Corr4

        jota = PARAM[0]*np.power(np.cos(theta),2)+PARAM[1]*np.cos(theta)+PARAM[2]+Corr

        Jotas.append(jota)

    Jotas = pd.DataFrame(Jotas)
    return Jotas

def error_calculator(calc, exp):
    return np.abs(calc - exp)

def J_error_calculator(isomers_jotas, J_exp):
    '''This function let you obtain the errors en scalar couplings while order
    the  J values which are interchangeable
    It returns a list of errors and a DataFrame of the calculated J ordered'''
    try:
        Jotas = J_exp.iloc[:, 0].copy()
        exchange = J_exp.iloc[:, 1].copy()

        calc_copy = isomers_jotas.copy()
        indices_intercambiables = [i for i in range(len(exchange)) if exchange[i] ==1]
        for e in indices_intercambiables:
            for i in range(isomers_jotas.shape[1]):
                calc_copy.iloc[e, i] = max([isomers_jotas.iloc[e, i], isomers_jotas.iloc[e+1, i]])
                calc_copy.iloc[e+1, i] = min([isomers_jotas.iloc[e, i], isomers_jotas.iloc[e+1, i]])
        return np.abs(calc_copy - pd.DataFrame(Jotas)), calc_copy, Jotas
    except:
        return isomers_jotas - J_exp, isomers_jotas, J_exp

def uf_DP4_calculator(scEC, scEH, scEJ, cant_comp):
    '''Function that calculate the DP4 probability'''
    t_distC = stats.t(11)
    t_distH = stats.t(14)
    t_distJ = stats.t(3)

    if len(scEC) > 0:
        pC=1-t_distC.cdf(scEC/2.306)
    else:
        pC = np.array(0)
    if len(scEH) > 0:
        pH=1-t_distH.cdf(scEH/0.185)
    else:
        pH = np.array(0)
    if len(scEJ) > 0:
        pJ=1-t_distJ.cdf(scEJ/0.992)
    else:
        pJ = np.array(0)
    try:
        sDP4H=100*np.prod(pH,0)/sum(np.prod(pH,0))
    except:
        sDP4H = np.array(['-' for _ in range(cant_comp)])
    try:
        sDP4C=100*np.prod(pC,0)/sum(np.prod(pC,0))
    except:
        sDP4C = np.array(['-' for _ in range(cant_comp)])
    try:
        sDP4J=100*np.prod(pJ,0)/sum(np.prod(pJ,0))
    except:
        sDP4J = np.array(['-' for _ in range(cant_comp)])
    try:
        sDP4full=100*np.prod(pH,0)*np.prod(pC,0)/sum(np.prod(pH,0)*np.prod(pC,0))
    except:
        sDP4full = np.array(['-' for _ in range(cant_comp)])
    try:
        DP4J=100*np.prod(pH,0)*np.prod(pC,0)*np.prod(pJ,0)/sum(np.prod(pH,0)*np.prod(pC,0)*np.prod(pJ,0))
    except:
        DP4J = np.array(['-' for _ in range(cant_comp)])
    return sDP4H, sDP4C, sDP4J, sDP4full, DP4J

def isomers_names(isomer_list, cant_comp):
    isomers = []
    for isom in isomer_list:
        isomer = 'Isomer '
        for caracter in isom:
            if caracter != '_':
                isomer += caracter
            else:
                break
        isomers.append(isomer)           
    return isomers

def modify_report(report):
    wb = load_workbook(report)
    ws = wb['DP4']
    ws.column_dimensions['A'].width = 20
    for i in range(2, 7):
        ws[f'A{i}'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    ws['A9'].font = Font(color = 'FF0000', bold = True)
    wb.save('Results_ML_J_DP4.xlsx')
    return

def main():
    print(f'In the directory {(Path(__file__).parent / "examples").as_posix()} you can find examples.')
    if change_directory('Select the input data directory.')==False:
        return
    # Reading data
    cant_comp, isomer_list = isomer_count()
    ds = data_sheet(cant_comp)
    if ds==False:
        return
    J_exp, Jwtl_all, d_exp_C, wtl_C_all, d_exp_H, wtl_H_all  = ds
    print("\nProcessing...")
    krC, krH = carga_kernels() 
    isomers_tensors_C = pd.DataFrame()
    isomers_tensors_H = pd.DataFrame()
    UnS_shifts_C = pd.DataFrame()
    UnS_shifts_H = pd.DataFrame()
    isomers_jotas = pd.DataFrame()
    isomers_shifts_C = pd.DataFrame()
    isomers_shifts_H = pd.DataFrame()
   


    for n_isom, isom in enumerate(isomer_list):

        conformers = glob.glob(f'{isom}*.log') +  glob.glob(f'{isom}*.out')
        conformers.sort() ##RG
        conformers.sort(key=lambda s: len(s)) #RG
        all_energies = []
        tens_all_conf_C = pd.DataFrame() 
        tens_all_conf_H = pd.DataFrame()
        jotas_all_conf = pd.DataFrame()
        wtl_C = label_check(wtl_C_all, n_isom+1)
        wtl_H = label_check(wtl_H_all, n_isom+1)
        Jwtl = label_check_J(Jwtl_all, n_isom+1)
        '''Preprocessing of energies to select only the conformers that will
        contribute to the Boltzmann, the limit is set at 3 kcal of the most stable
        '''
        for n,conf in enumerate(conformers):
            energy = get_energy(conf)
            all_energies.append(energy)
        e_relativas = relative_energies(all_energies)
        conf_energy = dict(zip(conformers, e_relativas)) 
        

        energies=[] 
        
        nuevo_isom=True #RG so we define the distance matrix only once

        for n,conf in enumerate(conformers):
            if conf_energy[conf] < 3:
                '''The energy, the tensors and a series of indices that serve
                to construct matrices are extracted from each conformer.
                '''
                lines, energy, tensors, start_coord, end_coord, start_mat, end_mat, start_mull, start_nbo, end_nbo = get_information(conf)

                energies.append(energy)

                '''Obtaining the formatted matrices of distances, coordinates,
                mulliken and hybridizations by using the extracted index
                '''
                mol_size, coord_mat, dist_mat, mull_mat = matrix(lines, start_coord, end_coord, start_mat, end_mat, start_mull)
                hibridization_mat = hibridizations(lines, mol_size, start_nbo, end_nbo)

            else:
                continue

            
            if nuevo_isom: #RG
                mat_vecinos = buscar_vecinos(dist_mat, coord_mat) 
                mat_vecinos_H = buscar_vecinos_H(hibridization_mat) 
                Jwtl = J_label_converter(Jwtl, mat_vecinos_H)
                nuevo_isom=False #RG
                
            '''Calculation of molecular descriptors for each conformation'''
            descriptors_C, descriptors_H, index_C, index_H = descriptor_calculator(tensors, coord_mat, mull_mat, hibridization_mat, mat_vecinos, dist_mat)

            '''Correction of the calculated tensors by the kernels loaded'''
            tensors_corrected_C, tensors_corrected_H = tensors_correction(descriptors_C, descriptors_H, krC, krH)
            tensors_indexed_C, tensors_indexed_H = tensor_indexation(tensors_corrected_C, tensors_corrected_H, index_C, index_H, mol_size)

            '''Coupling constant calculation'''
            UnSc_Jotas = J_calculator(Jwtl, coord_mat, hibridization_mat)
            Jotas = (UnSc_Jotas + 0.1405)/ 0.9509 

            '''Obtaining a vector with Boltzmann-weighted corrected tensors'''

            tens_C, tens_H = tens_ordenados(tensors_indexed_C, tensors_indexed_H, wtl_C, wtl_H)
            tens_all_conf_C = pd.concat([tens_all_conf_C,tens_C],axis=1) 
            tens_all_conf_H = pd.concat([tens_all_conf_H,tens_H],axis=1)
            jotas_all_conf = pd.concat([jotas_all_conf,Jotas],axis=1)

        e_relativas = relative_energies(energies)
        tensores_ponderados_C, tensores_ponderados_H, jotas_ponderadas =  Boltzman(tens_all_conf_C, tens_all_conf_H, jotas_all_conf, e_relativas)

        '''Chemical shifts calculation and Scaling procedure '''
        Unscaled_shift_C, Unscaled_shift_H = d_calculation(tensores_ponderados_C, tensores_ponderados_H)
        Scaled_shift_C, exp_C = escalado_CyH(Unscaled_shift_C, d_exp_C)
        Scaled_shift_H, exp_H = escalado_CyH(Unscaled_shift_H, d_exp_H)

        '''Once the corrected tensors and scaled chemical shifts of an isomer
        are obtained, they are added to the DataFrame of all the isomers that
        will be correlated with the experimental data '''

        isomers_tensors_C = pd.concat([isomers_tensors_C,tensores_ponderados_C],axis=1)
        isomers_tensors_H = pd.concat([isomers_tensors_H,tensores_ponderados_H],axis=1)

        UnS_shifts_C = pd.concat([UnS_shifts_C, Unscaled_shift_C ],axis=1)
        UnS_shifts_H = pd.concat([UnS_shifts_H, Unscaled_shift_H ],axis=1)

        isomers_jotas = pd.concat([isomers_jotas,jotas_ponderadas],axis=1)
        isomers_shifts_C = pd.concat([isomers_shifts_C, Scaled_shift_C ],axis=1)
        isomers_shifts_H = pd.concat([isomers_shifts_H, Scaled_shift_H ],axis=1)

    '''Large Error verifier'''
    UscEC = error_calculator(UnS_shifts_C, exp_C)
    UscEH = error_calculator(UnS_shifts_H, exp_H)

    '''Calculation of the H, C and J errors'''
    scEC = error_calculator(isomers_shifts_C, exp_C)
    scEH = error_calculator(isomers_shifts_H, exp_H)
    scEJ, isomers_jotas, J_exp = J_error_calculator(isomers_jotas, J_exp)

    '''Calculation of the partial or complete ufDP4'''
    sDP4H, sDP4C, sDP4J, sDP4full, DP4J = uf_DP4_calculator(scEC, scEH, scEJ, cant_comp)

    "Data procesing to print data to an excel file"
    isomers = isomers_names(isomer_list, cant_comp) 
    DP4_results = pd.DataFrame([sDP4H, sDP4C, sDP4J, sDP4full, DP4J], index= ['DP4-H', 'DP4-C', 'DP4-J', 'DP4-HC', 'JDP4full'], columns=(isomers))

    vac= pd.DataFrame([''])

    exp_Data = pd.concat([exp_C, vac, exp_H, vac, J_exp], axis = 0)
    NMR_Data = pd.concat([isomers_shifts_C, vac, isomers_shifts_H, vac, isomers_jotas ],axis=0)
    NMR_data = pd.concat([exp_Data, NMR_Data], axis = 1)
    NMR_data.columns = ['exp']+isomers
    
    NMR_Data_UnSc = pd.concat([UnS_shifts_C, vac, UnS_shifts_H, vac, isomers_jotas ],axis=0)
    NMR_data_UnSc = pd.concat([exp_Data, NMR_Data_UnSc], axis = 1)
    NMR_data_UnSc.columns = ['exp']+isomers

    errors =  pd.concat([scEC,vac, scEH, vac, scEJ],axis=0)
    error_data = pd.concat([exp_Data, errors], axis = 1)
    error_data.columns = ['exp']+isomers
    
    errors_Unscaled =  pd.concat([UscEC,vac, UscEH, vac, scEJ],axis=0)
    error_data_Unscaled = pd.concat([exp_Data, errors_Unscaled], axis = 1)
    error_data_Unscaled.columns = ['exp']+isomers

    print('\nDone!')
    perm=False
    
    while not perm:
        perm=True
        try:
            
            directorio = os.getcwd()
            workbook = os.path.join(directorio,'Results_ML_J_DP4.xlsx')
            print(f'\n -> Writting output file Results_ML_J_DP4.xlsx in {directorio:s}.')
            with open(workbook, 'w') as f:
                f.write('output file')
        except:
            print("   -> Can't write. Please choose another directory for output file.")
            perm=False
            if change_directory('Select the output directory.')==False:
                print('Exit.')
                return
            
        
    with pd.ExcelWriter(workbook) as writer: 
        DP4_results.to_excel(writer,sheet_name='DP4', index=True,float_format="%.2f")
        NMR_data_UnSc.to_excel(writer,sheet_name='Shifts_Unsc',index=False,float_format="%.4f")
        NMR_data.to_excel(writer,sheet_name='Shifts',index=False,float_format="%.4f")
        error_data_Unscaled.to_excel(writer,sheet_name='Unsc_Errors',index=False,float_format="%.4f")
        error_data.to_excel(writer,sheet_name='Errors',index=False,float_format="%.4f")
        
        
        
    modify_report('Results_ML_J_DP4.xlsx')


if __name__=='__main__' :
    main()
